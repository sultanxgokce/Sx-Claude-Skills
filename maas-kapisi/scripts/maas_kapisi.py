#!/usr/bin/env python3
"""maas-kapisi — aylık bordro ↔ banka maaş dosyası kapısı + arşiv.

Para hareketi başlatan bir zincirin önündeki makine kontrolü. Sessiz-yanlış YASAK:
bir şey ayrıştırılamıyorsa KIRMIZI der, tahminle devam etmez.

Kullanım:  maas_kapisi.py <komut> [...]
  al <YYYY-MM> <dosya...>   sınıfla + NFC-normalize + sha256 + arşivle
  kontrol <YYYY-MM>         karşılaştırma motoru -> rapor (md + json)
  ozet <YYYY-MM>            gruba gidecek standart onay metni (ekrana)
  getir <YYYY-MM> [tur]     arşivden dosya + sha256 doğrulaması
  durum                     tüm aylar envanteri
"""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import shutil
import sys
import unicodedata
from dataclasses import dataclass, field, asdict
from pathlib import Path

KOK = Path(os.environ.get("MAAS_KOK", "/config/evraklar/IK/05_Maas"))
TURLER = ("bordro", "maas-yukle", "puantaj", "pdks")
UZANTI = {"bordro": ".pdf", "maas-yukle": ".xlsx", "puantaj": ".xlsx", "pdks": ".xlsx"}

# "bordroda var / bankada yok" kuralı. Sultan 7-kişiyi teyit edince "kirmizi" yapılır.
EKSIK_KISI_SEVIYE = os.environ.get("MAAS_EKSIK_KISI", "sari")

MON = r"-?\d{1,3}(?:\.\d{3})*,\d{2}"          # 1.234.567,89
AY_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")


# ───────────────────────────────────────────────────────────── para (tam sayı kuruş)
def tl2k(s: str) -> int:
    """'30.622,35' -> 3062235 kuruş. Float yuvarlama hatası olmaz."""
    return int(round(float(s.replace(".", "").replace(",", ".")) * 100))


def f2k(v) -> int:
    return int(round(float(v) * 100))


def k2tl(k: int) -> str:
    """3062235 -> '30.622,35'"""
    isaret = "-" if k < 0 else ""
    k = abs(k)
    return f"{isaret}{k // 100:,}".replace(",", ".") + f",{k % 100:02d}"


def maske(tckn: str) -> str:
    """KVKK: TCKN tam yazılmaz."""
    return "•••" + tckn[-4:] if len(tckn) >= 4 else "•••"


# ───────────────────────────────────────────────────────────────────── veri tipleri
@dataclass
class Kisi:
    tckn: str
    ad: str
    net_k: int                      # kuruş
    giris: str = ""
    cikis: str = ""
    tgun: str = ""


@dataclass
class Bulgu:
    seviye: str                     # kirmizi | sari | istisna
    kod: str
    mesaj: str
    detay: list = field(default_factory=list)
    anahtar: str = ""               # TCKN — yalnız eşleştirme için, RAPORA BASILMAZ
    tutar_k: int = 0                # istisnayı tutara BAĞLAR (tutar değişirse istisna düşer)


# ───────────────────────────────────────────────────────────────── BORDRO (PDF)
def bordro_oku(pdf: Path):
    """Sayfa başına bir(+) kişi. Net tutar İKİ bağımsız yolla çıkarılır; uyuşmazsa KIRMIZI."""
    from pypdf import PdfReader

    kisiler: dict[str, Kisi] = {}
    bulgular: list[Bulgu] = []
    ham: dict[str, str] = {}        # ad → kayıt satırının kalanı (gün kontrolü için)
    beyan_toplam_k = None
    kayit_re = re.compile(r"^\s*(\d{1,3})\s+(.+?)\s+(\d{11})\s+(.*)$")

    sayfalar = PdfReader(str(pdf)).pages
    for pi, sayfa in enumerate(sayfalar, 1):
        L = (sayfa.extract_text() or "").split("\n")

        # bu sayfadaki kayıt-satırı indeksleri (+ TOPLAM bloğu bir sınır sayılır)
        idx = [i for i, l in enumerate(L) if kayit_re.match(l)]
        toplam_i = next((i for i, l in enumerate(L) if l.strip().startswith("TOPLAM ")), None)
        sinirlar = idx + ([toplam_i] if toplam_i is not None else []) + [len(L)]

        for n, i in enumerate(idx):
            son = sinirlar[sinirlar.index(i) + 1] if i in sinirlar else len(L)
            m = kayit_re.match(L[i])
            sira, ad, tckn, kalan = m.group(1), m.group(2).strip(), m.group(3), m.group(4)

            # --- yöntem A: 'Net Kazanç' ankraji (bu kaydın sınırları içinde)
            a = None
            for j in range(i, son):
                if L[j].strip() == "Net Kazanç":
                    for k in range(j + 1, min(j + 4, son)):
                        mm = re.fullmatch(MON, L[k].strip())
                        if mm:
                            a = tl2k(mm.group(0))
                            break
                    break
            # --- yöntem B: kayıt devam satırının son parasal alanı
            b = None
            if i + 1 < son:
                nums = re.findall(MON, L[i + 1])
                if nums:
                    b = tl2k(nums[-1])

            if a is None or b is None or a != b:
                bulgular.append(Bulgu(
                    "kirmizi", "pdf-capraz-dogrulama",
                    f"s{pi} · {ad}: net tutar iki yoldan farklı çıktı "
                    f"(A={k2tl(a) if a is not None else '?'} · B={k2tl(b) if b is not None else '?'}) "
                    "— ayrıştırılamadı, tahminle devam edilmedi"))
                continue

            if tckn in kisiler:
                bulgular.append(Bulgu("kirmizi", "bordro-tekrar-tckn",
                                      f"{ad} · {maske(tckn)} bordroda birden fazla kez var"))
                continue

            tarihler = re.findall(r"\d{2}/\d{2}/\d{4}", kalan)
            gun = re.search(r"G\s+(\d+)\s+(\d+)\s", kalan)
            kisiler[tckn] = Kisi(tckn, ad, a,
                                 giris=tarihler[0] if tarihler else "",
                                 cikis=tarihler[1] if len(tarihler) > 1 else "",
                                 tgun=gun.group(1) if gun else "")
            ham[ad] = kalan

        # --- PDF'in kendi beyan ettiği GENEL TOPLAM (bağımsız üçüncü kapı)
        if toplam_i is not None and beyan_toplam_k is None:
            for j in range(toplam_i, len(L)):
                if L[j].strip() == "Net Kazanç":
                    for k in range(j + 1, min(j + 4, len(L))):
                        mm = re.fullmatch(MON, L[k].strip())
                        if mm:
                            beyan_toplam_k = tl2k(mm.group(0))
                            break
                    break

    if not kisiler:
        bulgular.append(Bulgu("kirmizi", "bordro-bos",
                              "bordrodan hiç kişi çıkarılamadı — biçim tanınmadı"))

    hesap_k = sum(k.net_k for k in kisiler.values())
    if beyan_toplam_k is None:
        bulgular.append(Bulgu("sari", "bordro-beyan-yok",
                              "bordroda GENEL TOPLAM satırı bulunamadı — kişi toplamı çapraz doğrulanamadı"))
    elif beyan_toplam_k != hesap_k:
        bulgular.append(Bulgu(
            "kirmizi", "bordro-ic-tutarsiz",
            f"bordronun beyan ettiği toplam {k2tl(beyan_toplam_k)} ≠ "
            f"kişilerin toplamı {k2tl(hesap_k)} (fark {k2tl(beyan_toplam_k - hesap_k)})"))

    meta = {"sayfa": len(sayfalar), "kisi": len(kisiler),
            "beyan_toplam": beyan_toplam_k, "hesap_toplam": hesap_k, "ham": ham}
    return kisiler, bulgular, meta


# ─────────────────────────────────────────────────────── MAAŞ YÜKLE (TGB xlsx)
def maas_oku(xlsx: Path):
    import openpyxl

    bulgular: list[Bulgu] = []
    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    ad = next((s for s in wb.sheetnames if "Maaş" in s or "Maas" in s), wb.sheetnames[0])
    ws = wb[ad]
    # 'Toplam Adet' / 'Toplam Tutar' bu şablonda FORMÜLDÜR. Önbellek değeri yoksa
    # data_only=True None döner; ham kopya formülü görüp durumu ayırt etmemizi sağlar.
    ws_ham = openpyxl.load_workbook(str(xlsx), data_only=False)[ad]

    def hucre(r, c):
        return ws.cell(r, c).value

    def beyan(r, c, etiket):
        """Beyan alanı okunamıyorsa SESSİZ ATLAMA YOK — kırmızı."""
        v = hucre(r, c)
        if v is not None and not isinstance(v, str):
            return v
        if isinstance(v, str) and v.strip():
            try:
                return float(v.replace(".", "").replace(",", "."))
            except ValueError:
                pass
        ham = ws_ham.cell(r, c).value
        if isinstance(ham, str) and ham.startswith("="):
            bulgular.append(Bulgu(
                "kirmizi", "beyan-hesaplanmamis",
                f"'{etiket}' bir FORMÜL ve hesaplanmış değeri yok ({ham}) — "
                "dosya Excel'de açılıp kaydedilmeli; bu haliyle iç tutarlılık doğrulanamaz"))
        else:
            bulgular.append(Bulgu("kirmizi", "beyan-yok",
                                  f"'{etiket}' okunamadı ({ham!r}) — çapraz kontrol yapılamaz"))
        return None

    b_adet = beyan(4, 2, "Toplam Adet")
    b_tutar = beyan(5, 2, "Toplam Tutar")
    meta = {
        "kurum_kodu": hucre(1, 2), "sube_kodu": hucre(2, 2), "hesap": hucre(3, 2),
        "beyan_adet": int(b_adet) if b_adet is not None else None,
        "beyan_tutar_k": f2k(b_tutar) if b_tutar is not None else None,
        "doviz": str(hucre(6, 2) or "").strip(), "odeme_tarihi": str(hucre(7, 2) or "").strip(),
        "odeme_tipi": str(hucre(8, 2) or "").strip(), "borc_izahat": str(hucre(9, 2) or "").strip(),
    }

    if str(hucre(12, 1) or "").strip() != "İsim":
        bulgular.append(Bulgu("kirmizi", "maas-bicim",
                              "beklenen başlık düzeni yok (A12 'İsim' değil) — dosya tanınmadı"))
        return {}, bulgular, meta

    kayitlar: dict[str, dict] = {}
    hesaplar: dict[str, str] = {}
    for r in range(13, ws.max_row + 1):
        isim = hucre(r, 1)
        if not isim:
            continue
        ham_tckn, tutar = hucre(r, 2), hucre(r, 7)
        tckn = str(int(ham_tckn)) if isinstance(ham_tckn, (int, float)) else str(ham_tckn or "").strip()
        isim = str(isim).strip()

        if not re.fullmatch(r"\d{11}", tckn):
            bulgular.append(Bulgu("kirmizi", "maas-tckn-gecersiz",
                                  f"satır {r} · {isim}: TCKN 11 hane değil"))
            continue
        if tutar is None:
            bulgular.append(Bulgu("kirmizi", "maas-tutar-yok", f"satır {r} · {isim}: tutar boş"))
            continue
        tk = f2k(tutar)
        if tk <= 0:
            bulgular.append(Bulgu("kirmizi", "maas-tutar-gecersiz",
                                  f"{isim}: tutar {k2tl(tk)} (sıfır/negatif)"))
        if tckn in kayitlar:
            bulgular.append(Bulgu("kirmizi", "maas-tekrar-tckn",
                                  f"{isim} · {maske(tckn)} banka dosyasında birden fazla satırda"))
            continue

        hes = f"{hucre(r, 4)}/{hucre(r, 5)}"
        if hes in hesaplar and str(hucre(r, 5) or "").strip():
            bulgular.append(Bulgu("kirmizi", "maas-tekrar-hesap",
                                  f"aynı hesap numarası iki kişide: {hesaplar[hes]} ve {isim}"))
        hesaplar[hes] = isim
        kayitlar[tckn] = {"ad": isim, "tutar_k": tk, "satir": r}

    # --- dosyanın kendi içi tutarlı mı
    top = sum(v["tutar_k"] for v in kayitlar.values())
    if meta["beyan_tutar_k"] is not None and meta["beyan_tutar_k"] != top:
        bulgular.append(Bulgu(
            "kirmizi", "maas-ic-tutarsiz",
            f"'Toplam Tutar' {k2tl(meta['beyan_tutar_k'])} ≠ satırların toplamı {k2tl(top)}"))
    if meta["beyan_adet"] is not None and int(meta["beyan_adet"]) != len(kayitlar):
        bulgular.append(Bulgu(
            "kirmizi", "maas-adet-tutarsiz",
            f"'Toplam Adet' {int(meta['beyan_adet'])} ≠ satır sayısı {len(kayitlar)}"))

    # --- başlık alanları
    if meta["odeme_tipi"] != "M":
        bulgular.append(Bulgu("kirmizi", "maas-odeme-tipi",
                              f"Ödeme Tipi 'M' (MAAŞ) değil: {meta['odeme_tipi']!r}"))
    t = meta["odeme_tarihi"]
    if not re.fullmatch(r"\d{8}", t):
        bulgular.append(Bulgu("kirmizi", "maas-tarih-bicim",
                              f"Ödeme Tarihi GGAAYYYY değil: {t!r}"))
    else:
        try:
            gun = dt.date(int(t[4:]), int(t[2:4]), int(t[:2]))
            if gun < dt.date.today():
                bulgular.append(Bulgu("kirmizi", "maas-tarih-gecmis",
                                      f"Ödeme Tarihi geçmişte: {gun:%d.%m.%Y}"))
        except ValueError:
            bulgular.append(Bulgu("kirmizi", "maas-tarih-gecersiz", f"Ödeme Tarihi geçersiz: {t!r}"))
    if meta["doviz"].upper() != "TL":
        bulgular.append(Bulgu("sari", "maas-doviz", f"Döviz Kodu 'TL' değil: {meta['doviz']!r}"))

    meta["hesap_toplam_k"] = top
    return kayitlar, bulgular, meta


# ────────────────────────────────────────────────────────────────── KARŞILAŞTIRMA
def karsilastir(bordro: dict[str, Kisi], maas: dict[str, dict], istisnalar: list | None = None):
    """Eşleştirme ADA GÖRE DEĞİL, TCKN'ye göre (yazım farkları ölçüldü).

    İstisnalar SUSTURMAZ — seviyeyi 'istisna'ya çeker ve raporda gerekçesiyle görünür.
    Bir istisna (kod, tckn, TUTAR) üçlüsüne bağlıdır: tutar değişirse istisna DÜŞER, kırmızı döner.
    """
    ist = istisnalar or []

    def istisna_bul(kod, tckn, tutar_k):
        for i in ist:
            if i["kod"] == kod and i["tckn"] == tckn and i.get("tutar_k", tutar_k) == tutar_k:
                return i
        return None

    b, bulgular = set(bordro), []

    # 1) tutar farkı — kuruş hassasiyeti
    fark = [(bordro[t].ad, bordro[t].net_k, maas[t]["tutar_k"], t)
            for t in b & set(maas) if bordro[t].net_k != maas[t]["tutar_k"]]
    for ad, bk, mk, t in sorted(fark, key=lambda x: -abs(x[1] - x[2])):
        i = istisna_bul("tutar-farki", t, bk - mk)
        bulgular.append(Bulgu(
            "istisna" if i else "kirmizi", "tutar-farki",
            f"{ad} · {maske(t)}: bordro {k2tl(bk)} ↔ banka {k2tl(mk)} (fark {k2tl(bk - mk)})"
            + (f" — ONAYLI: {i['gerekce']} ({i['onay']}, {i['ts'][:10]})" if i else ""),
            anahtar=t, tutar_k=bk - mk))

    # 2) bankada var / bordroda yok  → her zaman KIRMIZI
    for t in sorted(set(maas) - b, key=lambda t: -maas[t]["tutar_k"]):
        bulgular.append(Bulgu(
            "kirmizi", "bordroda-yok",
            f"{maas[t]['ad']} · {maske(t)} banka dosyasında var ama BORDRODA YOK "
            f"({k2tl(maas[t]['tutar_k'])})"))

    # 3) bordroda var / bankada yok → onaylananlar ayrılır, kalanlar uyarı olur
    eksik = sorted(b - set(maas), key=lambda t: -bordro[t].net_k)
    onayli = [(t, istisna_bul("bankada-yok", t, bordro[t].net_k)) for t in eksik]
    kalan = [t for t, i in onayli if i is None]
    gecen = [(t, i) for t, i in onayli if i is not None]

    def satir(t, i=None):
        d = {"ad": bordro[t].ad, "tckn": maske(t), "net": k2tl(bordro[t].net_k),
             "giris": bordro[t].giris, "cikis": bordro[t].cikis or "-", "tgun": bordro[t].tgun}
        if i:
            d["gerekce"] = i["gerekce"]
        return d

    if kalan:
        bulgular.append(Bulgu(
            EKSIK_KISI_SEVIYE, "bankada-yok",
            f"{len(kalan)} kişi bordroda var ama banka dosyasında yok "
            f"(toplam {k2tl(sum(bordro[t].net_k for t in kalan))}) — gerekçesi ONAYLANMAMIŞ",
            detay=[satir(t) for t in kalan]))
    if gecen:
        bulgular.append(Bulgu(
            "istisna", "bankada-yok",
            f"{len(gecen)} kişi banka dosyasında yok — gerekçesi onaylı "
            f"(toplam {k2tl(sum(bordro[t].net_k for t, _ in gecen))})",
            detay=[satir(t, i) for t, i in gecen]))

    # 4) aynı TCKN'de ad yazım farkı
    def sadeles(s):
        """Yalnız boşluk/noktalama gürültüsünü atar; TÜRKÇE HARF FARKINI KORUR.

        NFKD KULLANILMAZ: Ö'yü O+aksan'a ayırıp aksanı düşürürdü ve
        'GÖRGİS' ile 'GORGİS' aynı sanılırdı — rapor bu farkı göstermeli.
        """
        s = unicodedata.normalize("NFC", s).upper()
        return "".join(ch for ch in s if ch.isalpha())

    ad_fark = [(bordro[t].ad, maas[t]["ad"]) for t in b & set(maas)
               if sadeles(bordro[t].ad) != sadeles(maas[t]["ad"])]
    if ad_fark:
        bulgular.append(Bulgu(
            "sari", "ad-yazim-farki",
            f"{len(ad_fark)} kişide ad yazımı iki dosyada farklı (TCKN aynı — eşleşme sağlam)",
            detay=[{"bordro": a, "banka": m} for a, m in ad_fark]))

    ozet = {
        "bordro_kisi": len(bordro),
        "bordro_toplam": k2tl(sum(k.net_k for k in bordro.values())),
        "maas_kisi": len(maas),
        "maas_toplam": k2tl(sum(v["tutar_k"] for v in maas.values())),
        "maas_toplam_k": sum(v["tutar_k"] for v in maas.values()),
        "eksik_kisi": len(eksik),
        "eksik_toplam": k2tl(sum(bordro[t].net_k for t in eksik)),
        "tutar_farki_adet": len(fark),
        "tutar_farki_toplam": k2tl(sum(bk - mk for _, bk, mk, _ in fark)),
    }
    return bulgular, ozet


# ────────────────────────────────────────────────────────────────────── ARŞİV
def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for blok in iter(lambda: f.read(1 << 20), b""):
            h.update(blok)
    return h.hexdigest()


def sinifla(p: Path) -> str | None:
    """İçeriğe göre sınıflar — dosya ADINA GÜVENMEZ ('TEMMUZ.xlsx' hiçbir şey söylemiyor)."""
    u = p.suffix.lower()
    if u == ".pdf":
        try:
            from pypdf import PdfReader
            t = (PdfReader(str(p)).pages[0].extract_text() or "")
            return "bordro" if "BORDRO" in t.upper() else None
        except Exception:
            return None
    if u not in (".xlsx", ".xlsm"):
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
        adlar = " ".join(wb.sheetnames).upper()
        if "MAAŞ" in adlar or "MAAS" in adlar:
            return "maas-yukle"
        ws = wb[wb.sheetnames[0]]
        if "TEMİZLİK" in adlar or "PUANTAJ" in adlar:
            return "puantaj"
        # PDKS: gövdede saat değerleri (09:00:00)
        for r in ws.iter_rows(min_row=2, max_row=6, max_col=12, values_only=True):
            if any(isinstance(v, dt.time) or re.fullmatch(r"\d{2}:\d{2}:\d{2}", str(v or ""))
                   for v in r):
                return "pdks"
        return None
    except Exception:
        return None


def ay_dizin(ay: str) -> Path:
    return KOK / ay


def manifest_yolu(ay: str) -> Path:
    return ay_dizin(ay) / "_manifest.json"


def manifest_oku(ay: str) -> dict:
    p = manifest_yolu(ay)
    return json.loads(p.read_text("utf-8")) if p.exists() else {"ay": ay, "dosyalar": []}


def manifest_yaz(ay: str, m: dict) -> None:
    p = manifest_yolu(ay)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(m, ensure_ascii=False, indent=2), "utf-8")


def cmd_al(ay: str, dosyalar: list[str]) -> int:
    if not AY_RE.match(ay):
        print(f"HATA: ay 'YYYY-MM' olmalı, verilen: {ay!r}", file=sys.stderr)
        return 2
    man = manifest_oku(ay)
    var = {d["tur"] for d in man["dosyalar"]}
    rc = 0
    for ham in dosyalar:
        # macOS NFD tuzağı: ad NFC'ye normalize edilmezse dosya "yok" görünür
        p = Path(unicodedata.normalize("NFC", ham))
        if not p.exists():
            p = Path(ham)
        if not p.exists():
            aday = [x for x in Path(ham).parent.iterdir()
                    if unicodedata.normalize("NFC", x.name) ==
                    unicodedata.normalize("NFC", Path(ham).name)]
            if aday:
                p = aday[0]
        if not p.exists():
            print(f"🔴 bulunamadı: {ham}")
            rc = 1
            continue

        tur = sinifla(p)
        if tur is None:
            print(f"🟡 sınıflanamadı, atlandı: {p.name}  (elle bak)")
            rc = 1
            continue

        rev = ""
        if tur in var:
            n = 2
            while any(d["tur"] == tur and d.get("rev") == n for d in man["dosyalar"]):
                n += 1
            rev = f"_rev{n}"
        hedef = ay_dizin(ay) / tur / f"{tur}_{ay}{rev}{UZANTI[tur]}"
        hedef.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(p, hedef)                       # SİLME YOK — kopya
        h = sha256(hedef)
        man["dosyalar"].append({
            "tur": tur, "yol": str(hedef.relative_to(KOK)), "ozgun_ad": p.name,
            "sha256": h, "boyut": hedef.stat().st_size,
            "alim": dt.datetime.now().isoformat(timespec="seconds"),
            **({"rev": int(rev[4:])} if rev else {}),
        })
        var.add(tur)
        print(f"🟢 {tur:<11} ← {p.name}\n   → {hedef.relative_to(KOK)}  sha {h[:12]}…")
    manifest_yaz(ay, man)
    eksik = [t for t in TURLER if t not in var]
    if eksik:
        print(f"\n🟡 bu ayda henüz yok: {', '.join(eksik)}")
    return rc


def bul(ay: str, tur: str) -> Path | None:
    """Aynı türün en son revizyonunu döndürür."""
    kayitlar = [d for d in manifest_oku(ay)["dosyalar"] if d["tur"] == tur]
    if not kayitlar:
        return None
    son = max(kayitlar, key=lambda d: d.get("rev", 1))
    return KOK / son["yol"]


def cmd_getir(ay: str, tur: str | None) -> int:
    man = manifest_oku(ay)
    if not man["dosyalar"]:
        print(f"🔴 {ay} için arşivde kayıt yok")
        return 1
    hedefler = [d for d in man["dosyalar"] if tur is None or d["tur"] == tur]
    if not hedefler:
        print(f"🔴 {ay} · '{tur}' yok. Mevcut: {sorted({d['tur'] for d in man['dosyalar']})}")
        return 1
    rc = 0
    for d in hedefler:
        p = KOK / d["yol"]
        if not p.exists():
            print(f"🔴 {d['tur']}: manifest'te var ama DİSKTE YOK → {p}")
            rc = 1
            continue
        s = sha256(p)
        if s != d["sha256"]:
            print(f"🔴 {d['tur']}: sha256 UYUŞMUYOR — dosya değişmiş/bozulmuş\n   {p}")
            rc = 1
            continue
        print(f"🟢 {d['tur']:<11} {p}\n   özgün ad: {d['ozgun_ad']}  ·  alım: {d['alim']}  ·  sha ✓")
    return rc


def cmd_durum() -> int:
    if not KOK.exists():
        print(f"arşiv henüz yok: {KOK}")
        return 0
    aylar = sorted([p.name for p in KOK.iterdir() if p.is_dir() and AY_RE.match(p.name)],
                   reverse=True)
    if not aylar:
        print("arşivde ay yok")
        return 0
    print(f"{'AY':<9}" + "".join(f"{t:<12}" for t in TURLER) + f"{'KONTROL':<10}BANKA")
    print("-" * 78)
    for ay in aylar:
        man = manifest_oku(ay)
        v = {d["tur"] for d in man["dosyalar"]}
        rapor = ay_dizin(ay) / "_rapor" / f"kontrol_{ay}.json"
        durum = json.loads(rapor.read_text("utf-8")).get("sonuc", "?") if rapor.exists() else "—"
        tp = ay_dizin(ay) / "_teyit.json"
        teyit = json.loads(tp.read_text("utf-8")).get("sonuc", "?") if tp.exists() else "—"
        print(f"{ay:<9}" + "".join(f"{'✓' if t in v else '·':<12}" for t in TURLER)
              + f"{durum:<10}{teyit}")
    return 0


# ──────────────────────────────────────────────────────────────────── KONTROL
PUANTAJ_KOD = ["X", "HT", "DZ", "Öİ", "RÇ", "İİ", "Eİ", "Dİ",
               "Yİ", "RA", "Üİ", "HS", "NÖİ", "İA"]


def roster_oku() -> dict:
    p = KOK / "_index" / "roster.json"
    return json.loads(p.read_text("utf-8")) if p.exists() else {"alias": {}, "puantaj_disi": {}}


def puantaj_oku(xlsx: Path):
    """Puantaj sayfasından kişi → kod sayaçları. TCKN YOK — yalnız isim var."""
    import openpyxl
    ws = openpyxl.load_workbook(str(xlsx), data_only=True)["TEMİZLİK"]
    if ws.cell(4, 35).value != "ÇALIŞILAN GÜN":
        return None, [Bulgu("kirmizi", "puantaj-bicim",
                            "puantaj özet blogu beklenen yerde değil (AI4 'ÇALIŞILAN GÜN' değil)")]
    # Özet sütunları (AI..AV) COUNTIF FORMÜLÜDÜR. Önbellek değeri yoksa data_only=True
    # None döner ve naif `or 0` bunu SIFIR sayar → herkesin izni 0 görünür, sessiz-yanlış.
    # (Banka dosyasındaki 'beyan-hesaplanmamis' ile aynı sınıf hata.)
    ws_ham = openpyxl.load_workbook(str(xlsx), data_only=False)["TEMİZLİK"]
    kisiler, hesapsiz = {}, []
    for r in range(7, 45):
        ad = ws.cell(r, 2).value
        if not ad:
            continue
        sayac = {}
        for i, kod in enumerate(PUANTAJ_KOD):
            v, hamv = ws.cell(r, 35 + i).value, ws_ham.cell(r, 35 + i).value
            if v is None and isinstance(hamv, str) and hamv.startswith("="):
                hesapsiz.append(str(ad).strip())
                break
            sayac[kod] = int(v or 0)
        else:
            kisiler[str(ad).strip()] = sayac
    if hesapsiz:
        return None, [Bulgu(
            "kirmizi", "puantaj-hesaplanmamis",
            f"puantaj özet sütunları FORMÜL ve hesaplanmış değeri yok "
            f"({len(set(hesapsiz))} kişide) — dosya Excel'de açılıp kaydedilmeli; "
            "bu haliyle gün zinciri doğrulanamaz (sıfır sayılırdı)")]
    return kisiler, []


def gun_kontrol(puantaj: dict, bordro_ham: dict[str, str], bordro: dict[str, Kisi]):
    """Puantaj → bordro zinciri. Bu olmadan KOCA BİR HATA SINIFI görünmezdi:
    muhasebeci bir kişinin gününü yanlış girse bordro ve banka dosyası birbiriyle
    mükemmel uyuşur, kontrol TEMİZ der, kimse fark etmez.

    Ölçülmüş değişmez (2026-07, 38/38): puantaj (Üİ + DZ) == bordro 'Eksik Gün'.
    """
    r = roster_oku()
    alias, disi = r.get("alias", {}), r.get("puantaj_disi", {})
    bulgular = []

    def n(s):
        return "".join(c for c in unicodedata.normalize("NFC", s).upper() if c.isalpha())

    ham = {n(k): (k, v) for k, v in bordro_ham.items()}
    eslesen, toplam_ui = 0, 0

    for pad, kod in puantaj.items():
        bad = alias.get(pad, pad)
        if n(bad) not in ham:
            bulgular.append(Bulgu(
                "kirmizi", "puantaj-eslesmedi",
                f"{pad}: puantajda var, bordroda karşılığı YOK "
                f"(yazım farkıysa roster.json'a alias ekleyin — tahminle eşleştirilmez)"))
            continue
        eslesen += 1
        gercek_ad, kalan = ham[n(bad)]
        m = re.match(r"^\s*(\d+)\s*Gün\s+(\d+)\s", kalan)
        eksik_b = int(m.group(1)) if m else 0
        kod_b = m.group(2) if m else "-"
        eksik_p = kod["Üİ"] + kod["DZ"]
        toplam_ui += kod["Üİ"]
        if eksik_b != eksik_p:
            # DZ yolu 2026-07'de HİÇ oluşmadı → o dala 'ölçülmedi' etiketi
            olculdu = kod["DZ"] == 0
            bulgular.append(Bulgu(
                "kirmizi" if olculdu else "sari", "eksik-gun-farki",
                f"{gercek_ad}: puantaj Üİ={kod['Üİ']}+DZ={kod['DZ']} (={eksik_p}) ↔ "
                f"bordro Eksik Gün={eksik_b} (kod {kod_b})"
                + ("" if olculdu else " — DZ dalı hiç ölçülmedi, sarı bırakıldı")))

    kapsam = {n(k) for k in disi}
    kayip = [bordro[t].ad for t in bordro
             if n(bordro[t].ad) not in {n(alias.get(p, p)) for p in puantaj}
             and n(bordro[t].ad) not in kapsam]
    if kayip:
        bulgular.append(Bulgu(
            "sari", "puantajda-yok",
            f"{len(kayip)} kişi bordroda var, puantajda yok ve kapsam-dışı listesinde de değil",
            detay=[{"ad": a} for a in kayip]))

    return bulgular, {"eslesen": eslesen, "puantaj_kisi": len(puantaj),
                      "toplam_ucretsiz_izin": toplam_ui}


def istisna_yolu(ay: str) -> Path:
    return ay_dizin(ay) / "_istisna.json"


def istisna_oku(ay: str) -> list:
    p = istisna_yolu(ay)
    return json.loads(p.read_text("utf-8")) if p.exists() else []


def cmd_istisna(ay: str, kod: str, kisi: str, gerekce: str, onay: str) -> int:
    """Sultan'ın onayladığı meşru sapmayı AYA ÖZEL kaydeder.

    Ay klasöründe durur → sonraki ay TEMİZ başlar, aynı sapma yine kırmızı yakar.
    Bu bilerekdir: 'hesabı yok' geçici bir durumdur, sessizce kalıcılaşmamalı.
    """
    if kod not in ("tutar-farki", "bankada-yok"):
        print("🔴 kod: tutar-farki | bankada-yok")
        return 2
    bp, mp = bul(ay, "bordro"), bul(ay, "maas-yukle")
    if bp is None or mp is None:
        print(f"🔴 {ay}: bordro/maas-yukle eksik")
        return 1
    bordro, _, _ = bordro_oku(bp)
    maas, _, _ = maas_oku(mp)

    hedef = [t for t in bordro if kisi.upper() in bordro[t].ad.upper()] if kisi != "hepsi" else None
    if kod == "bankada-yok":
        aday = [t for t in bordro if t not in maas]
    else:
        aday = [t for t in set(bordro) & set(maas) if bordro[t].net_k != maas[t]["tutar_k"]]
    if hedef is not None:
        aday = [t for t in aday if t in hedef]
    if not aday:
        print(f"🔴 '{kisi}' için '{kod}' bulgusu YOK — var olmayan bir şeye istisna yazılmaz")
        return 1

    kayit = istisna_oku(ay)
    for t in aday:
        tk = bordro[t].net_k if kod == "bankada-yok" else bordro[t].net_k - maas[t]["tutar_k"]
        kayit = [k for k in kayit if not (k["kod"] == kod and k["tckn"] == t)]
        kayit.append({"kod": kod, "tckn": t, "ad": bordro[t].ad, "tutar_k": tk,
                      "gerekce": gerekce, "onay": onay,
                      "ts": dt.datetime.now().isoformat(timespec="seconds")})
        print(f"🟢 istisna: {bordro[t].ad} · {kod} · {k2tl(tk)} — {gerekce} ({onay})")
    istisna_yolu(ay).parent.mkdir(parents=True, exist_ok=True)
    istisna_yolu(ay).write_text(json.dumps(kayit, ensure_ascii=False, indent=2), "utf-8")
    print(f"\n→ {len(kayit)} istisna kayıtlı. `kontrol {ay}` ile yeniden değerlendir.")
    return 0


def cmd_kontrol(ay: str) -> int:
    bp, mp = bul(ay, "bordro"), bul(ay, "maas-yukle")
    if bp is None or mp is None:
        print(f"🔴 {ay}: kontrol için ikisi de gerekli — "
              f"bordro {'✓' if bp else 'YOK'} · maas-yukle {'✓' if mp else 'YOK'}")
        return 1

    bordro, b_bulgu, b_meta = bordro_oku(bp)
    maas, m_bulgu, m_meta = maas_oku(mp)
    c_bulgu, ozet = karsilastir(bordro, maas, istisna_oku(ay))
    bulgular = b_bulgu + m_bulgu + c_bulgu

    # --- puantaj → bordro zinciri (puantaj arşivdeyse)
    pp, g_ozet = bul(ay, "puantaj"), None
    if pp is None:
        bulgular.append(Bulgu("sari", "puantaj-yok",
                              "puantaj arşivde yok — gün zinciri doğrulanamadı"))
    else:
        pua, p_bulgu = puantaj_oku(pp)
        bulgular += p_bulgu
        if pua:
            g_bulgu, g_ozet = gun_kontrol(pua, b_meta["ham"], bordro)
            bulgular += g_bulgu

    kirmizi = [x for x in bulgular if x.seviye == "kirmizi"]
    sari = [x for x in bulgular if x.seviye == "sari"]
    istisna = [x for x in bulgular if x.seviye == "istisna"]
    sonuc = "KIRMIZI" if kirmizi else ("SARI" if sari else "TEMIZ")

    S = [f"# Maaş kontrol raporu — {ay}", "",
         f"**Sonuç: {sonuc}**  ·  {len(kirmizi)} kırmızı · {len(sari)} sarı",
         f"_üretim: {dt.datetime.now():%d.%m.%Y %H:%M}_", "",
         "## Kaynaklar", "",
         f"- bordro: `{bp.name}` — {b_meta['sayfa']} sayfa, {b_meta['kisi']} kişi",
         f"- banka : `{mp.name}` — {len(maas)} satır, ödeme {m_meta['odeme_tarihi']}, "
         f"tip {m_meta['odeme_tipi']}", "",
         "## Sayılar", "",
         "| | kişi | toplam |", "|---|---:|---:|",
         f"| bordro (net ödenen) | {ozet['bordro_kisi']} | {ozet['bordro_toplam']} |",
         f"| banka maaş dosyası | {ozet['maas_kisi']} | {ozet['maas_toplam']} |",
         f"| **fark** | **{ozet['bordro_kisi'] - ozet['maas_kisi']}** | "
         f"**{k2tl(sum(k.net_k for k in bordro.values()) - ozet['maas_toplam_k'])}** |", ""]

    if g_ozet:
        S += ["### Puantaj → bordro zinciri", "",
              f"- puantajdaki {g_ozet['puantaj_kisi']} kişinin {g_ozet['eslesen']}'i "
              f"bordroda eşleşti",
              f"- ücretsiz izin toplamı: **{g_ozet['toplam_ucretsiz_izin']} gün** "
              "(puantaj Üİ+DZ ↔ bordro *Eksik Gün*)", ""]

    if ozet["eksik_kisi"] or ozet["tutar_farki_adet"]:
        S += ["### Farkın kapanışı", "",
              f"- bankada olmayan {ozet['eksik_kisi']} kişi: {ozet['eksik_toplam']}",
              f"- tutar farkları ({ozet['tutar_farki_adet']} kişi): {ozet['tutar_farki_toplam']}", ""]

    for baslik, liste in (("🔴 Kırmızı — onay akışı başlamaz", kirmizi),
                          ("🟡 Sarı — teyit gerekir", sari),
                          ("✅ Onaylı istisna — gerekçesi kayıtlı, YİNE DE her ay teyit edilir", istisna)):
        if not liste:
            continue
        S += [f"## {baslik}", ""]
        for x in liste:
            S.append(f"- **[{x.kod}]** {x.mesaj}")
            if x.detay:
                S.append("")
                anahtar = list(x.detay[0].keys())
                S.append("  | " + " | ".join(anahtar) + " |")
                S.append("  |" + "|".join(["---"] * len(anahtar)) + "|")
                for d in x.detay:
                    S.append("  | " + " | ".join(str(d[a]) for a in anahtar) + " |")
            S.append("")
    if sonuc == "TEMIZ":
        S += ["## ✅ Tüm kapılar temiz", "",
              "Bordro ile banka dosyası kişi kişi, kuruşu kuruşuna örtüşüyor.", ""]

    rd = ay_dizin(ay) / "_rapor"
    rd.mkdir(parents=True, exist_ok=True)
    (rd / f"kontrol_{ay}.md").write_text("\n".join(S), "utf-8")
    (rd / f"kontrol_{ay}.json").write_text(json.dumps({
        "ay": ay, "sonuc": sonuc, "uretim": dt.datetime.now().isoformat(timespec="seconds"),
        "ozet": ozet, "bordro_meta": b_meta,
        "maas_meta": {k: v for k, v in m_meta.items() if k != "beyan_tutar_k"},
        "bulgular": [asdict(x) for x in bulgular],
    }, ensure_ascii=False, indent=2), "utf-8")

    print("\n".join(S))
    print(f"\n→ rapor: {rd / f'kontrol_{ay}.md'}")
    return 2 if kirmizi else 0


def cmd_ozet(ay: str) -> int:
    p = ay_dizin(ay) / "_rapor" / f"kontrol_{ay}.json"
    if not p.exists():
        print(f"🔴 önce `kontrol {ay}` çalıştır")
        return 1
    r = json.loads(p.read_text("utf-8"))
    if r["sonuc"] == "KIRMIZI":
        print(f"🔴 {ay} KIRMIZI — onay metni ÜRETİLMEZ. Önce bulgular çözülmeli.")
        return 2
    o = r["ozet"]
    print("──────── gruba gidecek metin ────────")
    print(f"{o['maas_kisi']} sayıda işçiye {o['maas_toplam']} TL maaş ödemesi Garanti Bankası "
          "otomatik maaş yükle ile gönderilecektir. Onaylıyor musunuz Fahri Bey?")
    print("─────────────────────────────────────")
    if r["sonuc"] == "SARI":
        print(f"\n🟡 not: raporda teyit bekleyen sarı bulgu var — göndermeden önce bak.")
    return 0


WA = "/config/.claude/skills/whatsapp-gonder/scripts/wa-gonder.sh"
GRUP = os.environ.get("MAAS_GRUP", "BORDRO & MAAŞ")


def _onay_metni(ay: str):
    p = ay_dizin(ay) / "_rapor" / f"kontrol_{ay}.json"
    if not p.exists():
        return None, f"🔴 önce `kontrol {ay}` çalıştır"
    r = json.loads(p.read_text("utf-8"))
    if r["sonuc"] == "KIRMIZI":
        return None, f"🔴 {ay} KIRMIZI — gönderim YOK. Önce bulgular çözülmeli."
    o = r["ozet"]
    return (f"{o['maas_kisi']} sayıda işçiye {o['maas_toplam']} TL maaş ödemesi Garanti Bankası "
            "otomatik maaş yükle ile gönderilecektir. Onaylıyor musunuz Fahri Bey?"), None


FEDERE = "/config/.claude/skills/federe-os-cekirdek/scripts/federe.sh"
# Banka dosyasının gitmesi GEREKEN şirket hesabı. Farklıysa teyit KIRMIZI verir.
GONDERICI = os.environ.get("MAAS_GONDERICI", "fahrigokce@gmail.com")


def _ad_esit(a: str, b: str) -> bool:
    """Banka teyidi dosya adını bozuk kodlamayla döndürüyor
    ('Fahri Go¨kc¸e 7.Ay Maas¸.xlsx'). Aksan/boşluk/noktalama atılıp karşılaştırılır."""
    def sad(s):
        s = unicodedata.normalize("NFKD", s).lower()
        return "".join(c for c in s if c.isalnum())
    return sad(a) == sad(b)


def teyit_ayristir(metin: str) -> dict:
    """Garanti BBVA 'Yükleme Bildirimi' e-postasından alanları çeker."""
    d = {}
    for anahtar, kalip in (
        ("dosya", r"Dosya\s*Ad[ıi]\s*:?\s*(.+?)\s*$"),
        ("boyut", r"Dosya\s*Boyutu\s*:?\s*([\d.]+)\s*Byte"),
        ("kurum", r"Kurum\s*:?\s*(\d+)"),
        ("tarih", r"Ula[şs]ma\s*Zaman[ıi]\s*:?\s*([\d.]{8,10})"),
        ("kanal", r"Geli[şs]\s*Kanal[ıi]\s*:?\s*(\S+@\S+)"),
    ):
        m = re.search(kalip, metin, re.M | re.I)
        if m:
            d[anahtar] = m.group(1).strip()
    if "boyut" in d:
        d["boyut"] = int(d["boyut"].replace(".", ""))
    d["basarili"] = bool(re.search(r"Ba[şs]ar[ıi]l[ıi]\s*Y[üu]kleme", metin, re.I))
    return d


def cmd_teyit(ay: str, a) -> int:
    """Bankanın 'aldım' dediği dosya, kontrolden GEÇEN dosya mı?

    Bu kapı olmadan zincir kontrolden sonra KOPUYORDU: doğrulama arşivdeki dosyaya,
    gönderim ise insanın elindeki dosyaya bakıyordu. İkisinin aynı olduğunu kimse ölçmüyordu.
    """
    mp = bul(ay, "maas-yukle")
    if mp is None:
        print(f"🔴 {ay}: arşivde maas-yukle yok — neyle karşılaştıracağımı bilmiyorum")
        return 1

    if a.metin_dosya:
        p = Path(a.metin_dosya)
        if not p.exists():
            print(f"🔴 bulunamadı: {p}")
            return 1
        t = teyit_ayristir(p.read_text("utf-8", errors="replace"))
    else:
        t = {"dosya": a.dosya, "boyut": a.boyut, "kurum": a.kurum,
             "tarih": a.tarih, "kanal": a.kanal, "basarili": True}
    t = {k: v for k, v in t.items() if v is not None}

    if not t.get("boyut"):
        print("🔴 teyitten 'Dosya Boyutu' okunamadı — karşılaştırma yapılamaz "
              "(--metin-dosya ya da --boyut ver)")
        return 2

    import openpyxl
    kurum_dosyada = openpyxl.load_workbook(str(mp), data_only=True)[
        "TGB Yeni Maaş Dosyası"].cell(1, 2).value
    kayit = [d for d in manifest_oku(ay)["dosyalar"] if d["tur"] == "maas-yukle"]
    ozgun = max(kayit, key=lambda d: d.get("rev", 1))["ozgun_ad"]
    boyut = mp.stat().st_size

    kirmizi, yesil = [], []

    def kapi(ok, iyi, kotu):
        (yesil if ok else kirmizi).append(iyi if ok else kotu)

    kapi(t.get("basarili", False), "banka 'Başarılı Yükleme' dedi",
         "banka BAŞARILI demiyor — yükleme geçmemiş olabilir")
    kapi(t["boyut"] == boyut,
         f"dosya boyutu birebir ({boyut} byte)",
         f"BOYUT FARKLI — bankaya giden {t['boyut']} byte, kontrolden geçen {boyut} byte")
    if t.get("dosya"):
        kapi(_ad_esit(t["dosya"], ozgun),
             f"dosya adı örtüşüyor ({ozgun})",
             f"AD FARKLI — banka: {t['dosya']!r} · arşiv: {ozgun!r}")
    if t.get("kurum"):
        kapi(str(t["kurum"]) == str(kurum_dosyada),
             f"kurum kodu doğru ({kurum_dosyada})",
             f"KURUM KODU FARKLI — banka: {t['kurum']} · dosya: {kurum_dosyada}")
    if t.get("kanal"):
        kapi(t["kanal"].lower() == GONDERICI.lower(),
             f"doğru hesaptan gönderilmiş ({t['kanal']})",
             f"YANLIŞ HESAPTAN gönderilmiş — {t['kanal']} (olması gereken: {GONDERICI})")

    kanit_yol = None
    if a.kanit:
        k = Path(unicodedata.normalize("NFC", a.kanit))
        if k.exists():
            hedef = ay_dizin(ay) / "banka-teyit" / f"banka-teyit_{ay}{k.suffix}"
            hedef.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(k, hedef)
            kanit_yol = str(hedef.relative_to(KOK))
        else:
            print(f"🟡 kanıt bulunamadı, atlandı: {a.kanit}")

    sonuc = "KIRMIZI" if kirmizi else "TEYITLI"
    (ay_dizin(ay) / "_teyit.json").write_text(json.dumps({
        "ay": ay, "sonuc": sonuc, "banka": t, "arsiv": {"ozgun_ad": ozgun, "boyut": boyut,
                                                        "sha256": sha256(mp)},
        "kanit": kanit_yol, "ts": dt.datetime.now().isoformat(timespec="seconds"),
    }, ensure_ascii=False, indent=2), "utf-8")

    print(f"── BANKA TEYİDİ · {ay} · {sonuc} ──")
    for s in yesil:
        print(f"  ✅ {s}")
    for s in kirmizi:
        print(f"  🔴 {s}")
    if kanit_yol:
        print(f"  📎 kanıt arşivlendi: {kanit_yol}")
    print("\n⚠️  Dürüst sınır: banka özet (hash) vermiyor. Ad+boyut+kurum tutuyorsa bu "
          "GÜÇLÜ bir tutarlılık kanıtıdır, mutlak kimlik ispatı değildir.")
    return 2 if kirmizi else 0


def cmd_bildir(ay: str, merkeze: bool, kime: str) -> int:
    """Onay metnini üretir. GRUBA DOĞRUDAN YAZMAZ — merkez basar.

    Merkez usulü (s01 talimatı 2026-08-07): 'BORDRO & MAAŞ' maaş ödemesini tetikleyen
    insan-onay kapısıdır. Yazma yetkisi bir kutuya verilirse o kutudaki HER ajan ve HER
    cron para zincirine metin düşürebilir. Bu yüzden mmex'in izin listesi bilerek yalnız
    'Sultan'; doğrudan gönderim yolu buradan KALDIRILDI (disiplin değil, mekanik).
    """
    import subprocess
    metin, hata = _onay_metni(ay)
    if hata:
        print(hata)
        return 2
    if not merkeze:
        print(f"── ONAY METNİ · {ay} ──\n{metin}\n──\n"
              f"Gruba merkez basar:  bildir {ay} --merkeze\n"
              f"(Sultan elle de yapıştırabilir. Bu kutudan gruba doğrudan yazılmaz.)")
        return 0
    if not Path(FEDERE).exists():
        print(f"🔴 federe istemcisi yok: {FEDERE}")
        return 3
    baslik = f"MMEx/NAMIK: {ay} maas onay metni HAZIR - gruba basilmasi rica olunur"
    r = subprocess.run(["bash", FEDERE, "gonder", "s01", baslik[:120], "maas-kapisi",
                        metin[:500]], capture_output=True, text=True)
    print(r.stdout.strip() or r.stderr.strip())
    if r.returncode != 0:
        print(f"🔴 merkeze bırakılamadı (çıkış {r.returncode}) — METİN GİTMEDİ")
        return r.returncode
    kayit = ay_dizin(ay) / "_rapor" / f"bildirim_{ay}.json"
    kayit.write_text(json.dumps({"ay": ay, "yol": "merkez(s01)", "metin": metin,
                                 "ts": dt.datetime.now().isoformat(timespec="seconds")},
                                ensure_ascii=False, indent=2), "utf-8")
    print("\n⚠️  Bırakıldı ≠ basıldı. Merkez gruba basınca sonucu bildirecek.")
    return 0


def cmd_mail_taslak(ay: str) -> int:
    """Bankaya gidecek e-postanın İÇERİĞİNİ üretir. GÖNDERME YOK — bilerek yok."""
    metin, hata = _onay_metni(ay)
    if hata:
        print(hata)
        return 2
    ek = bul(ay, "maas-yukle")
    r = json.loads((ay_dizin(ay) / "_rapor" / f"kontrol_{ay}.json").read_text("utf-8"))
    o, m = r["ozet"], r["maas_meta"]
    yil, ayn = ay.split("-")
    aylar = ["", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz",
             "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
    print(f"""── E-POSTA TASLAĞI · {ay} ──  (bu komut e-posta GÖNDERMEZ)

Kimden : fahrigokce@gmail.com
Kime   : <Garanti maaş yükle adresi — Sultan doğrulayacak>
Konu   : Fahri Gökçe Elektronik İnş. Ltd. Şti. — {aylar[int(ayn)]} {yil} Maaş Ödemesi
Ek     : {ek}
         (kurum {m['kurum_kodu']} · şube {m['sube_kodu']} · hesap {m['hesap']})

Gövde:
Merhaba,

Ekte {aylar[int(ayn)]} {yil} dönemine ait maaş ödeme dosyamız yer almaktadır.
{o['maas_kisi']} personel için toplam {o['maas_toplam']} TL ödeme yapılacaktır.
Ödeme tarihi: {m['odeme_tarihi'][:2]}.{m['odeme_tarihi'][2:4]}.{m['odeme_tarihi'][4:]}

Bilgilerinize arz ederiz.
Saygılarımızla,
Fahri Gökçe Elektronik İnş. Ltd. Şti.
──
🔒 Gönderme adımı bilerek bu skill'de YOK: banka maili para talimatıdır, insanda kalır.""")
    return 0


PCLOUD = "/config/.claude/skills/pcloud-erisim/scripts/pcloud.sh"
IKPC = "/config/.claude/skills/ik-arsiv/scripts/ikpc.sh"
PCLOUD_KOK = os.environ.get("MAAS_PCLOUD_KOK", "25200500122")   # MMEx-IK-Arsiv


def cmd_yedekle(ay: str) -> int:
    """Ayın arşivini pCloud'a aynalar. Hub tek nokta değil (2 cihaza senkron) ama
    dış, sürümlü yedek yoktu — bu onu kapatır. Mevcut skill'ler BESTELENİR."""
    import subprocess
    d = ay_dizin(ay)
    if not d.exists():
        print(f"🔴 {ay} arşivde yok")
        return 1
    for s in (PCLOUD, IKPC):
        if not Path(s).exists():
            print(f"🔴 gerekli skill yok: {s}")
            return 3

    dosyalar = sorted(p for p in d.rglob("*") if p.is_file())
    rc, yuklenen = 0, 0
    for p in dosyalar:
        rel = p.relative_to(KOK).parent
        mk = subprocess.run(["bash", IKPC, "mkpath", PCLOUD_KOK, f"05_Maas/{rel}"],
                            capture_output=True, text=True)
        fid = (mk.stdout or "").strip().split()[-1] if mk.returncode == 0 else ""
        if not fid.isdigit():
            print(f"🔴 klasör açılamadı: 05_Maas/{rel}  ({(mk.stderr or mk.stdout).strip()[:80]})")
            rc = 1
            continue
        up = subprocess.run(["bash", PCLOUD, "upload", str(p), fid],
                            capture_output=True, text=True)
        if up.returncode == 0:
            yuklenen += 1
            print(f"  🟢 {p.relative_to(KOK)}")
        else:
            print(f"  🔴 {p.relative_to(KOK)} — {(up.stderr or up.stdout).strip()[:80]}")
            rc = 1
    print(f"\n{yuklenen}/{len(dosyalar)} dosya pCloud'a yüklendi"
          + ("" if rc == 0 else "  — EKSİK VAR, yedek TAM DEĞİL"))
    return rc


def main() -> int:
    ap = argparse.ArgumentParser(prog="maas-kapisi", description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="komut", required=True)
    p = sub.add_parser("al"); p.add_argument("ay"); p.add_argument("dosyalar", nargs="+")
    for k in ("kontrol", "ozet"):
        sub.add_parser(k).add_argument("ay")
    p = sub.add_parser("getir"); p.add_argument("ay"); p.add_argument("tur", nargs="?")
    p = sub.add_parser("istisna"); p.add_argument("ay")
    p.add_argument("kod", choices=["tutar-farki", "bankada-yok"])
    p.add_argument("kisi", help="ad parçası ya da 'hepsi'")
    p.add_argument("gerekce")
    p.add_argument("--onay", default="Sultan")
    p = sub.add_parser("bildir"); p.add_argument("ay")
    p.add_argument("--merkeze", action="store_true",
                   help="onay metnini merkeze (s01) bırak — gruba MERKEZ basar")
    p.add_argument("--kime", default=GRUP, help=argparse.SUPPRESS)
    sub.add_parser("mail-taslak").add_argument("ay")
    p = sub.add_parser("teyit", help="bankanın aldığı dosya, kontrolden geçen dosya mı?")
    p.add_argument("ay")
    p.add_argument("--metin-dosya", dest="metin_dosya",
                   help="banka teyit e-postasının metni (kopyala-yapıştır dosyası)")
    p.add_argument("--dosya"); p.add_argument("--boyut", type=int)
    p.add_argument("--kurum"); p.add_argument("--tarih"); p.add_argument("--kanal")
    p.add_argument("--kanit", help="ekran görüntüsü/pdf — arşive kopyalanır")
    sub.add_parser("yedekle").add_argument("ay")
    sub.add_parser("durum")
    a = ap.parse_args()

    if a.komut == "al":
        return cmd_al(a.ay, a.dosyalar)
    if a.komut == "kontrol":
        return cmd_kontrol(a.ay)
    if a.komut == "ozet":
        return cmd_ozet(a.ay)
    if a.komut == "getir":
        return cmd_getir(a.ay, a.tur)
    if a.komut == "istisna":
        return cmd_istisna(a.ay, a.kod, a.kisi, a.gerekce, a.onay)
    if a.komut == "bildir":
        return cmd_bildir(a.ay, a.merkeze, a.kime)
    if a.komut == "mail-taslak":
        return cmd_mail_taslak(a.ay)
    if a.komut == "teyit":
        return cmd_teyit(a.ay, a)
    if a.komut == "yedekle":
        return cmd_yedekle(a.ay)
    if a.komut == "durum":
        return cmd_durum()
    return 2


if __name__ == "__main__":
    sys.exit(main())
