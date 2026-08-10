#!/usr/bin/env python3
"""MUTASYON TESTİ — motorun gerçekten yakaladığını kanıtlar.

Temiz bir taban üretilir (Temmuz banka dosyasından Diyaddin düzeltilerek), sonra
tek tek kusur enjekte edilir. Her enjeksiyon BEKLENEN kodu KIRMIZI vermezse test kalır.
Ayrıca temiz taban KIRMIZI verirse (yalancı-pozitif) test kalır.
"""
from __future__ import annotations

import os
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))

import openpyxl                                    # noqa: E402
import maas_kapisi as M                            # noqa: E402

SS = Path("/config/projects/MMEx/ss")
ARK = Path(os.environ["MAAS_KOK"])
AY = "2026-07"
gecti, kaldi = 0, []


def bank_yukle(p: Path):
    wb = openpyxl.load_workbook(str(p))
    return wb, wb["TGB Yeni Maaş Dosyası"]


def kos(bank: Path, bordro: Path):
    b, bb, _ = M.bordro_oku(bordro)
    m, mb, _ = M.maas_oku(bank)
    c, _ = M.karsilastir(b, m)
    return [x for x in bb + mb + c if x.seviye == "kirmizi"]


def kontrol(ad: str, kirmizilar, bekle: str | None):
    global gecti
    kodlar = {x.kod for x in kirmizilar}
    if bekle is None:
        ok = not kirmizilar
        neden = f"beklenmedik kırmızı: {sorted(kodlar)}"
    else:
        ok = bekle in kodlar
        neden = f"'{bekle}' çıkmadı; çıkanlar: {sorted(kodlar) or 'HİÇBİRİ'}"
    if ok:
        gecti += 1
        print(f"  GECTI  {ad}")
    else:
        kaldi.append(ad)
        print(f"  KALDI  {ad}  → {neden}")


bordro = ARK / AY / "bordro" / f"bordro_{AY}.pdf"
kaynak = ARK / AY / "maas-yukle" / f"maas-yukle_{AY}.xlsx"
tmp = Path(tempfile.mkdtemp())

# ── TABAN: Diyaddin'i bordroya eşitle + 7 eksik kişiyi 'sari' bırak → KIRMIZI olmamalı
taban = tmp / "taban.xlsx"
shutil.copy2(kaynak, taban)
wb, ws = bank_yukle(taban)
for r in range(13, 60):
    if ws.cell(r, 1).value and "DİYADDİN" in str(ws.cell(r, 1).value):
        ws.cell(r, 7).value = 17968.32
        break
# B4/B5 şablonda FORMÜLDÜR; taban'da sabit değere çevrilir ki enjeksiyonlar
# 'beyan-hesaplanmamis' gürültüsü yerine hedefledikleri kapıyı sınasın.
dolu = [r for r in range(13, 60) if ws.cell(r, 1).value]
ws.cell(4, 2).value = len(dolu)
ws.cell(5, 2).value = round(sum(ws.cell(r, 7).value or 0 for r in dolu), 2)
wb.save(taban)

print("=" * 66)
print("0) TABAN — düzeltilmiş dosya KIRMIZI vermemeli (yalancı-pozitif kapısı)")
print("=" * 66)
kontrol("temiz taban", kos(taban, bordro), None)

MUT = [
    ("a) bir tutarı 1 KURUŞ oynat", "tutar-farki",
     lambda ws: (ws.cell(14, 7), setattr_(ws.cell(14, 7), round(ws.cell(14, 7).value + 0.01, 2)))),
    # Satır silmek HEM adedi HEM toplamı bozar; taban B4/B5'i sabitlediği için
    # ikisi de ateşlemeli. Adet kapısının ayrıca canlı olduğu (i) ile sınanır.
    ("b) bir satırı SİL", "maas-adet-tutarsiz",
     lambda ws: ws.delete_rows(20, 1)),
    ("i) 'Toplam Adet'i tek başına boz", "maas-adet-tutarsiz",
     lambda ws: setattr_(ws.cell(4, 2), 99)),
    ("j) beyan alanını FORMÜLE çevir (önbelleksiz)", "beyan-hesaplanmamis",
     lambda ws: setattr_(ws.cell(5, 2), "=SUM(G13:G60)")),
    ("c) 'Toplam Tutar'ı boz", "maas-ic-tutarsiz",
     lambda ws: setattr_(ws.cell(5, 2), 999999.99)),
    ("d) iki kişiye AYNI hesap no", "maas-tekrar-hesap",
     lambda ws: (setattr_(ws.cell(14, 4), ws.cell(13, 4).value),
                 setattr_(ws.cell(14, 5), ws.cell(13, 5).value))),
    ("e) 'Ödeme Tipi'ni boz", "maas-odeme-tipi",
     lambda ws: setattr_(ws.cell(8, 2), "X")),
    ("f) bordroda OLMAYAN kişi ekle", "bordroda-yok",
     lambda ws: [setattr_(ws.cell(60, c), v) for c, v in
                 ((1, "HAYALET KİŞİ"), (2, 12345678901), (3, "62"), (4, 512),
                  (5, "9999999"), (7, 5000.0))]),
    ("g) tutarı SIFIR yap", "maas-tutar-gecersiz",
     lambda ws: setattr_(ws.cell(15, 7), 0.0)),
    ("h) 'Ödeme Tarihi'ni geçmişe al", "maas-tarih-gecmis",
     lambda ws: setattr_(ws.cell(7, 2), "01012020")),
]


def setattr_(hucre, deger):
    hucre.value = deger
    return None


print()
print("=" * 66)
print("1) ENJEKSİYONLAR — her biri KIRMIZI vermeli")
print("=" * 66)
for ad, bekle, boz in MUT:
    p = tmp / f"mut_{bekle}_{abs(hash(ad))}.xlsx"
    shutil.copy2(taban, p)
    wb, ws = bank_yukle(p)
    boz(ws)
    wb.save(p)
    kontrol(f"{ad}  [{bekle}]", kos(p, bordro), bekle)

# ── PDF tarafı: bozuk/ilgisiz PDF sessizce yanlış sonuç vermemeli
print()
print("=" * 66)
print("2) BORDRO PDF — bozuk girdi sessiz-yanlış vermemeli")
print("=" * 66)
sahte = tmp / "sahte.pdf"
sahte.write_bytes(b"%PDF-1.4\n% bu bir bordro degil\n")
try:
    b, bul, _ = M.bordro_oku(sahte)
    ok = not b and any(x.seviye == "kirmizi" for x in bul)
    print(f"  {'GECTI' if ok else 'KALDI'}  bordro olmayan PDF → kişi yok + kırmızı")
    gecti += ok or kaldi.append("sahte pdf")
except Exception as e:
    gecti += 1
    print(f"  GECTI  bordro olmayan PDF → dürüst hata ({type(e).__name__})")

# ── Geriye dönük külliyat: eski .xls sessiz-yanlış vermemeli
print()
print("=" * 66)
print("3) ESKİ .xls KÜLLİYATI — sessiz-yanlış yerine dürüst hata")
print("=" * 66)
eski = Path("/config/evraklar/Fahri/Pc/Word-Excel")
for ad in ("Maas-07-02-2020.xls", "HAZİRAN_bordro.xls"):
    p = eski / ad
    if not p.exists():
        continue
    try:
        k, bul, _ = M.maas_oku(p)
        ok = not k or any(x.seviye == "kirmizi" for x in bul)
        print(f"  {'GECTI' if ok else 'KALDI'}  {ad} → {'reddedildi' if ok else 'SESSİZ KABUL!'}")
        gecti += ok or kaldi.append(ad)
    except Exception as e:
        gecti += 1
        print(f"  GECTI  {ad} → dürüst hata ({type(e).__name__})")

# ── PII sızıntısı
print()
print("=" * 66)
print("4) PII SIZINTISI — üretilen hiçbir metinde tam TCKN olmamalı")
print("=" * 66)
import re                                          # noqa: E402
b, bb, _ = M.bordro_oku(bordro)
m, mb, _ = M.maas_oku(kaynak)
c, ozet = M.karsilastir(b, m)
metin = " ".join([x.mesaj + str(x.detay) for x in bb + mb + c] + [str(ozet)])
gercek = {t for t in list(b) + list(m)}
sizan = [t for t in gercek if t in metin]
ok = not sizan
print(f"  {'GECTI' if ok else 'KALDI'}  bulgu metinlerinde tam TCKN: {len(sizan)} adet")
gecti += ok or kaldi.append("PII sızıntısı")

# ── BANKA TEYİDİ: her kapı gerçekten ateşliyor mu?
print()
print("=" * 66)
print("5) BANKA TEYİDİ — bankanın aldığı dosya, kontrolden geçen dosya mı?")
print("=" * 66)
import argparse                                    # noqa: E402
import contextlib                                  # noqa: E402
import io                                          # noqa: E402

arsiv = M.bul(AY, "maas-yukle")
GERCEK = dict(boyut=arsiv.stat().st_size, kurum="673885",
              ad="Fahri Go¨kc¸e 7.Ay Maas¸.xlsx", kanal="fahrigokce@gmail.com",
              basarili="Başarılı Yükleme")


def teyit_metni(**ov):
    d = {**GERCEK, **ov}
    return (f"MAAŞ Ödeme Dosyası Yükleme Bildirimi - {d['basarili']} Kurum:{d['kurum']}\n"
            f" Dosya Adı    : {d['ad']}\n"
            f" Dosya Boyutu : {d['boyut']} Byte\n"
            f" Garanti BBVA` ya Ulaşma Zamanı : 7.08.2026\n"
            f" Geliş Kanalı : {d['kanal']}\n")


def teyit_kos(**ov):
    p = tmp / f"teyit_{abs(hash(str(ov)))}.txt"
    p.write_text(teyit_metni(**ov), "utf-8")
    ns = argparse.Namespace(metin_dosya=str(p), dosya=None, boyut=None, kurum=None,
                            tarih=None, kanal=None, kanit=None)
    with contextlib.redirect_stdout(io.StringIO()):
        return M.cmd_teyit(AY, ns)


for ad, bekle, ov in [
    ("gercek teyit (yalanci-pozitif kapisi)", 0, {}),
    ("boyut 1 byte oynatilmis", 2, {"boyut": GERCEK["boyut"] + 1}),
    ("kurum kodu farkli (yanlis kuruma yuklenmis)", 2, {"kurum": "999999"}),
    ("YANLIS HESAPTAN gonderilmis", 2, {"kanal": "baskasi@gmail.com"}),
    ("dosya adi farkli (eski kopya)", 2, {"ad": "Fahri Gokce 6.Ay Maas.xlsx"}),
    ("banka BASARILI demiyor", 2, {"basarili": "Hatali Yukleme"}),
]:
    rc = teyit_kos(**ov)
    ok = rc == bekle
    (globals().__setitem__("gecti", gecti + 1) if ok else kaldi.append(f"teyit: {ad}"))
    print(f"  {'GECTI' if ok else 'KALDI'}  {ad}  (rc={rc}, beklenen={bekle})")

# teyit dosyasini gercek haline geri getir (arsiv kirlenmesin)
teyit_kos()

# ── PUANTAJ → BORDRO zinciri
print()
print("=" * 66)
print("6) PUANTAJ → BORDRO — muhasebecinin gün hatası yakalanıyor mu?")
print("=" * 66)
pua_yol = M.bul(AY, "puantaj")
b_kisi, _, b_meta = M.bordro_oku(bordro)


def gun_kos(**boz):
    """Puantaj kopyasını bozup gün kontrolünü koşar; kırmızı kodları döndürür.

    ÖNEMLİ: openpyxl ile kaydetmek COUNTIF önbelleğini SİLER. Bu yüzden mutasyondan önce
    özet sütunları SABİT değere çevrilir — yoksa her test 'puantaj-hesaplanmamis'
    verir ve hedeflenen kapı hiç sınanmamış olur (test kendi kendini kandırırdı).
    """
    if not boz:
        pua, hata = M.puantaj_oku(pua_yol)          # bozulmamış: orijinali kullan
    else:
        p = tmp / f"pua_{abs(hash(str(boz)))}.xlsx"
        shutil.copy2(pua_yol, p)
        say = openpyxl.load_workbook(str(p), data_only=True)["TEMİZLİK"]
        deger = {(r, c): say.cell(r, c).value
                 for r in range(7, 45) for c in range(35, 49)}
        wb = openpyxl.load_workbook(str(p))
        ws = wb["TEMİZLİK"]
        for (r, c), v in deger.items():
            ws.cell(r, c).value = v                  # formül → sabit (önbellek korunur)
        for r, c, v in boz.get("hucre", []):
            ws.cell(r, c).value = v
        wb.save(p)
        pua, hata = M.puantaj_oku(p)
    if pua is None:
        return {x.kod for x in hata if x.seviye == "kirmizi"}
    bul_, _ = M.gun_kontrol(pua, b_meta["ham"], b_kisi)
    return {x.kod for x in bul_ if x.seviye == "kirmizi"}


# AR sutunu (45) = Üİ sayaci; satir 7 = ilk personel
for ad, bekle, boz in [
    ("gercek puantaj (yalanci-pozitif kapisi)", None, {}),
    ("bir kisinin ucretsiz izni 1 gun ARTIRILDI", "eksik-gun-farki",
     {"hucre": [(7, 45, 1)]}),
    ("bir kisinin ADI degistirildi (roster'da yok)", "puantaj-eslesmedi",
     {"hucre": [(7, 2, "AHMET YILMAZ")]}),
    ("ozet sutunlari ONBELLEKSIZ formule cevrildi", "puantaj-hesaplanmamis",
     {"hucre": [(7, 45, "=COUNTIF(D7:AH7,\"Üİ\")")]}),
]:
    kodlar = gun_kos(**boz)
    ok = (not kodlar) if bekle is None else (bekle in kodlar)
    (globals().__setitem__("gecti", gecti + 1) if ok else kaldi.append(f"gun: {ad}"))
    print(f"  {'GECTI' if ok else 'KALDI'}  {ad}  → {sorted(kodlar) or 'kirmizi yok'}")

print()
print("=" * 66)
toplam = gecti + len(kaldi)
print(f"SONUC: {gecti}/{toplam} gecti")
if kaldi:
    print("KALANLAR:", kaldi)
    sys.exit(1)
print("MOTOR KABUL EDILDI — her enjeksiyon yakalandi, temiz taban yalanci-pozitif vermedi")
